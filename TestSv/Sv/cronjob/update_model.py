from datetime import datetime
import requests
from dotenv import load_dotenv
from pymongo import MongoClient
import os
import json
import pandas as pd
import pickle
from sklearn import linear_model
import openpyxl

def preprocess_city_name(city: str, lower=False):
        import re
        replace_for_province = ['Tỉnh', 'Thành phố']
        big_regex = re.compile('|'.join(map(re.escape, replace_for_province)))
        return big_regex.sub('', city).strip().lower() if lower else big_regex.sub('', city).strip()

def get_province_name_by_code(code: str):
        from vietnam_provinces.enums import ProvinceEnum
        return ProvinceEnum[f'P_{code}'].value.name

def add_new_data(new_data, data_path = None, result_path = None, workbook = None, sheet_name = 'Sheet1'):
    if new_data is not None:
        if workbook is None:
            workbook = openpyxl.load_workbook(filename=data_path)
        ws = workbook[sheet_name]
        row = ws.max_row + 1
        for col, entry in enumerate(new_data, start=1):
            ws.cell(row=row, column=col, value=entry)
        workbook.save(result_path)

def retrain_model(data_path, result_path):
    df = pd.read_excel(data_path).dropna().drop_duplicates()
    x_train = df.drop("Places", axis="columns")
    y_train = df.Places
    # new retrain model
    lm_model = linear_model.LinearRegression()
    lm_model.fit(x_train.values,y_train.values)
    with open(result_path, "wb") as f:
        pickle.dump(lm_model, f)

# load env
load_dotenv()
DOMAIN_NAME = os.getenv('DOMAIN_NAME')
SECRET_KEY = os.getenv('SECRET_KEY')
CONNECTION_STRING = os.getenv('CONNECTION_STRING')
DATA_PATH_LINEAR_MODEL = os.getenv('DATA_PATH_LINEAR_MODEL')
RESULT_PATH_LINEAR_MODEL = os.getenv('RESULT_PATH_LINEAR_MODEL')
DATA_PATH_DECISIONTREE_MODEL = os.getenv('DATA_PATH_DECISIONTREE_MODEL')
RESULT_PATH_DECESIONTREE_MODEL = os.getenv('RESULT_PATH_DECESIONTREE_MODEL')
headers = {'Authorization': SECRET_KEY}

client = MongoClient(CONNECTION_STRING)

log_tour_created = client.recommender.log_tour_created
# log_tour_created_today = list(log_tour_created.find({'log_created_date': {'$gte': datetime(2023,5,29)}}, {'_id': 0, 'log_created_date': 0}))
log_tour_created_today = list(log_tour_created.find({'log_created_date': {'$gte': datetime(datetime.now().year,datetime.now().month,datetime.now().day,0,0,0,0)}}, {'_id': 0, 'log_created_date': 0}))
if len(log_tour_created_today) > 0:
    linear_workbook = openpyxl.load_workbook(filename=DATA_PATH_LINEAR_MODEL)
    decisiontree_workook = openpyxl.load_workbook(filename=DATA_PATH_DECISIONTREE_MODEL)
    for log in log_tour_created_today:
        pois: list = log['pois']
        req = requests.post(DOMAIN_NAME+'api/v2/extract_info_to_excels', headers=headers, data=json.dumps(log)).json()
        if len(req['error']) == 0:
            # data for linear model
            a_linear_row = req['data']
            places = 0
            # data for decision tree model
            # from(str) to(str) num_of_day(int) price(float) transport(str)
            a_decision_tree_row = [preprocess_city_name(get_province_name_by_code(log['from'][0])),
                                preprocess_city_name(get_province_name_by_code(log['to'][0])),
                                a_linear_row[0][0],
                                a_linear_row[0][2],
                                a_linear_row[0][-1]]
            # remove transport label from linear row
            for row in a_linear_row:
                row.pop(-1)


            for i in range(len(a_linear_row)):
                temp = []
                num_of_day = a_linear_row[i][0]
                pois_in_province = pois[:num_of_day]
                pois = [poi for poi in pois if poi not in pois_in_province]
                for poi in pois_in_province:
                    temp.extend(poi.split(','))
                temp = len(temp)
                a_linear_row[i].append(temp)

            # add new data to workbook
            [add_new_data(new_data=row, workbook=linear_workbook, result_path=DATA_PATH_LINEAR_MODEL) for row in a_linear_row]
            add_new_data(new_data=a_decision_tree_row, workbook=decisiontree_workook, result_path=RESULT_PATH_DECESIONTREE_MODEL)

    # after adding new data to workbook, retrain model with new data
    retrain_model(data_path=DATA_PATH_LINEAR_MODEL, result_path=RESULT_PATH_LINEAR_MODEL)

print('predict_n_places model updated at {} with {} tour created.'.format(datetime.now().strftime("%m/%d/%Y, %H:%M:%S"), len(log_tour_created_today)))
